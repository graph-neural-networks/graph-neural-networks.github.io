import dataclasses
from collections import defaultdict
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any
from typing import List, Dict
import xml.etree.ElementTree as ET
from pybtex import database

import pandas as pd

from pylatexenc.latex2text import LatexNodes2Text

import pandas as pd
import pytz
import ruamel
from ftfy import fix_text
from openpyxl import load_workbook
from pylatexenc.latex2text import LatexNodes2Text
from ruamel import yaml
from ruamel import yaml

# https://docs.google.com/spreadsheets/d/19LRnJpae5NQd0D1NEO40kTbwDvS9f125tpsjBdevrcs/edit#gid=0
from scripts.dataentry.paths import *

fix = {
    "490": "WS-21",
    "510": "WS-1",
    "884": "WS-13",
    "1093": "WS-11",
    "1761": "WS-4",
    "2217": "WS-17",
    "2487": "WS-16",
    "2575": "WS-3",
    "2797": "WS-13",
    "2800": "WS-13",
    "2976": "WS-25",
    "3476": "WS-6",
    "3561": "WS-25",
}


@dataclass
class Session:
    name: str
    start: datetime
    end: datetime
    host: str


@dataclass
class Workshop:
    uid: str
    sessions: List[Session]
    description: str


@dataclass
class Paper:
    uid: str
    ws_id: str
    title: str
    authors: str
    abstract: str
    track: str
    kind: str
    link: str


def load_workshop_overview_excel() -> pd.DataFrame:
    wb = load_workbook(PATH_WORKSHOPS_OVERVIEW)
    ws = wb.worksheets[0]
    ws.delete_rows(1, 1)
    ws.delete_rows(27, 100)
    ws.delete_cols(7, 3)
    ws.delete_cols(8, 14)

    emnlp_workshops = pd.read_csv(PATH_WORKSHOPS_CSV)

    softconf_id_to_organizers = {
        row["softconfNumber"]: row["authors"] for _, row in emnlp_workshops.iterrows()
    }

    df = pd.DataFrame(
        ws.values,
        columns=[
            "Softconf Number",
            "UID",
            "Name",
            "Summary",
            "Authors",
            "URL",
            "Alias",
            "Old UID",
        ],
    )
    df = df.dropna(subset=["UID"])
    df["Softconf Number"] = df["Softconf Number"].fillna(-1)

    df["Softconf Number"] = df["Softconf Number"].apply(lambda x: int(x))
    df["Organizers"] = df["Softconf Number"].apply(
        lambda x: softconf_id_to_organizers[x]
    )

    return df


def build_workshops_basics() -> List[Dict[str, Any]]:
    workshops = load_workshop_overview_excel()
    schedule = load_schedule()
    zooms = get_zooms()

    data = []
    for _, row in workshops.iterrows():
        uid = row["UID"].strip()
        if uid == "WS-22":
            continue

        alias = row["Alias"]

        if alias is None:
            other = {"WS-4": "SCAI", "WS-1": "ConLL", "WS-13": "DeeLIO"}
            alias = other[row["UID"]]

        workshop = schedule[uid]
        alias = alias.lower()
        sessions = [
            {
                "start_time": session.start,
                "end_time": session.end,
                "name": session.name,
                "hosts": session.host,
            }
            for session in workshop.sessions
        ]
        title = row["Name"].strip()
        entry = {
            "UID": uid,
            "title": title,
            "organizers": row["Organizers"].strip(),
            "abstract": workshop.description if workshop.description else row["Summary"],
            "website": row["URL"],
            "rocketchat_channel": f"workshop-{alias.lower()}",
            "alias": alias,
            "sessions": sessions,
        }

        if title in zooms:
            entry["zoom_links"] = zooms[title]
        else:

            # print(title)
            pass

        data.append(entry)

    data.sort(key=lambda w: -int(w["UID"][2:]))

    return data


def load_schedule() -> Dict[str, Workshop]:
    wb = load_workbook(PATH_WORKSHOPS_SCHEDULE)

    data = {}
    for ws in wb.worksheets[4:]:
        workshop_id = ws["B2"].value
        assert workshop_id.startswith("WS-"), "Does not start with WS: " + workshop_id

        description = ws["B3"].value or ""
        ws.delete_rows(1, 6)
        ws.delete_cols(7, 100)
        df = pd.DataFrame(
            ws.values,
            columns=[
                "Session Name",
                "Day",
                "Start Time",
                "End Time",
                "Time Zone",
                "Host",
            ],
        )
        df.dropna(subset=["Session Name"], inplace=True)

        sessions = []
        for idx, row in df.iterrows():
            name = row["Session Name"].strip()
            host = row["Host"] or "TBD"

            day = row["Day"]
            start_time = row["Start Time"]
            end_time = row["End Time"]
            tz_name = row["Time Zone"]

            if not name or not tz_name:
                continue

            if isinstance(start_time, str):
                start_time = datetime.strptime(start_time, "%H:%M").time()
            if isinstance(end_time, str):
                end_time = datetime.strptime(end_time, "%H:%M").time()

            if isinstance(start_time, datetime):
                start_time = start_time.time()
            if isinstance(end_time, datetime):
                end_time = end_time.time()

            tz = pytz.timezone(tz_name)

            start = datetime.combine(day.date(), start_time)
            start = tz.localize(start)

            if start_time > end_time:
                day += timedelta(days=1)

            end = datetime.combine(day.date(), end_time)
            end = tz.localize(end)

            session = Session(name, start, end, host)
            sessions.append(session)

        workshop = Workshop(workshop_id, sessions, description)
        assert workshop_id not in data, (
            "workshop id already in data",
            workshop_id,
            data[workshop_id],
        )
        data[workshop_id] = workshop

    return data


def load_slideslive():
    # https://docs.google.com/spreadsheets/d/1Cp04DGRiDj8oY00-xDjTpjzCd_fjq3YhqOclhvFRK94/edit#gid=0
    df = pd.read_csv(PATH_SLIDESLIVE_WORKSHOPS)
    df = df.drop([0])

    df_obj = df.select_dtypes(["object"])
    df[df_obj.columns] = df_obj.apply(lambda x: x.str.strip())

    workshop_df = load_workshop_overview_excel()

    ws_name_to_id = {
        row["Name"]: row["UID"].strip() for _, row in workshop_df.iterrows()
    }

    return df


def generate_workshop_papers(slideslive: pd.DataFrame):
    venues = []
    UIDs = []
    titles = []
    authors = []
    presentation_ids = []

    fixed_papers = {}

    i = 0
    for _, row in slideslive.iterrows():
        if is_not_paper(row):
            continue

        title = row["Title"].replace("\n", " ")
        title = LatexNodes2Text().latex_to_text(title)
        title = fix_text(title).strip()
        author_list = [
            fix_text(e.strip()) for e in re.split(",| and | And ", row["Speakers"])
        ]

        presentation_id = row["SlidesLive link"].replace("https://slideslive.com/", "")

        ws = row["Organizer track name"].strip()
        uid = row["Unique ID"].strip()

        if title.startswith("Findings:"):
            continue

        if uid == "510" or uid == "561" or uid == "1093" or uid == "1761"  or uid == "2575-ws3":
            continue

        if ws == "WS-15" and str(uid) in fix.keys():
            ws = fix[uid]
            fixed_papers[title] = i
        elif title in fixed_papers:
            idx = fixed_papers[title]
            # print(title, idx, len(titles))
            assert titles[idx] == title
            presentation_ids[idx] = presentation_id
            continue

        paper_id = f"{ws}.{uid}"

        venues.append(ws)
        UIDs.append(paper_id)
        titles.append(title)
        authors.append("|".join(author_list))
        presentation_ids.append(
            presentation_id
        )

        i += 1

    anthology_papers = get_anthology_workshop_papers() + read_wmt_bib()
    title_to_anthology_paper = {a.title.strip().lower(): a for a in anthology_papers}
    author_to_anthology_paper = {a.authors.lower(): a for a in anthology_papers}
    url_to_anthology_paper = {a.link: a for a in anthology_papers}

    unmatched = []
    uid_to_anthology_paper = {}
    for uid, title, author in zip(UIDs, titles, authors):
        if title.lower() in title_to_anthology_paper:
            assert uid not in uid_to_anthology_paper
            uid_to_anthology_paper[uid] = title_to_anthology_paper[title.lower()]
        elif uid == "WS-1.Shared8":
            uid_to_anthology_paper[uid] = url_to_anthology_paper[
                "https://www.aclweb.org/anthology/2020.conll-shared.1"
            ]
        else:
            unmatched.append((uid, title, author.lower()))

    for uid, title, author in list(unmatched):
        if author.lower() in author_to_anthology_paper:
            assert uid not in uid_to_anthology_paper, (
                uid,
                title,
                author,
                uid_to_anthology_paper[uid],
            )
            uid_to_anthology_paper[uid] = author_to_anthology_paper[author.lower()]
            unmatched.remove((uid, title, author.lower()))

    unmatched_df = pd.DataFrame(unmatched)
    unmatched_df.to_csv("yamls/unmatched_workshop_papers.csv", index=False)

    # print("Unmatched", len(unmatched), len(uid_to_anthology_paper))

    abstracts = []
    urls = []
    for i, uid in enumerate(UIDs):
        if uid in uid_to_anthology_paper:
            paper = uid_to_anthology_paper[uid]
            abstracts.append(paper.abstract)
            authors[i] = paper.authors
            titles[i] = paper.title
            urls.append(paper.link)
        else:
            abstracts.append("")
            urls.append("")

    all_the_titles = set(titles)
    not_slideslive_but_anthology = []

    for paper in anthology_papers:
        if (
            paper.title not in all_the_titles
            and paper.ws_id != "findings"
            and not paper.uid.startswith("2020.nlpcovid19-acl")
        ):
            venues.append(paper.ws_id)
            UIDs.append(f"{paper.ws_id}.{paper.uid}")
            titles.append(paper.title)
            authors.append(paper.authors)
            abstracts.append(paper.abstract)
            presentation_ids.append("")
            urls.append(paper.link)

    not_slideslive_but_anthology_df = pd.DataFrame(not_slideslive_but_anthology)
    not_slideslive_but_anthology_df.to_csv(
        "yamls/not_slideslive_but_anthology.csv", index=False
    )

    data = {
        "workshop": venues,
        "UID": UIDs,
        "title": titles,
        "authors": authors,
        "abstract": abstracts,
        "presentation_id": presentation_ids,
        "pdf_url": urls,
    }

    columns = [
        "workshop",
        "UID",
        "title",
        "authors",
        "abstract",
        "presentation_id",
        "pdf_url",
    ]
    df = pd.DataFrame(data, columns=columns)

    df.to_csv(PATH_YAMLS / "workshop_papers.csv", index=False)


def get_anthology_workshop_papers() -> List[Paper]:
    anthology = (
        Path(
            r"C:\Users\klie\AppData\Roaming\JetBrains\PyCharm2020.2\scratches\emnlp\acl-anthology"
        )
        / "data"
    )

    conference = "emnlp"
    year = 2020

    mapping = {
        "2020.conll-1": "WS-1",
        "2020.alw-1": "WS-17",
        "2020.blackboxnlp-1": "WS-25",
        "2020.clinicalnlp-1": "WS-12",
        "2020.cmcl-1": "WS-5",
        "2020.codi-1": "WS-16",
        "2020.deelio-1": "WS-13",
        "2020.eval4nlp-1": "WS-20",
        "2020.insights-1": "WS-3",
        "2020.intexsempar-1": "WS-6",
        "2020.louhi-1": "WS-19",
        "2020.nlpbt-1": "WS-23",
        "2020.nlpcovid19-1": "WS-26",
        "2020.nlpcss-1": "WS-18",
        "2020.nlposs-1": "WS-9",
        "2020.privatenlp-1": "WS-24",
        "2020.scai-1": "WS-4",
        "2020.sdp-1": "WS-7",
        "2020.sigtyp-1": "WS-11",
        "2020.splu-1": "WS-10",
        "2020.spnlp-1": "WS-21",
        "2020.sustainlp-1": "WS-15",
        "2020.wnut-1": "WS-14",
        "2020.findings-1": "findings",
    }

    papers = []
    for venue, ws_id in mapping.items():
        if venue.endswith("-1"):
            file_name = venue[:-2]
        else:
            file_name = venue

        path_to_xml = anthology / "xml" / f"{file_name}.xml"
        tree = ET.parse(path_to_xml)
        root = tree.getroot()
        collection_id = root.attrib["id"]

        for volume in root.findall("volume"):

            volume_id = volume.attrib["id"]

            for paper in volume.findall("paper"):
                paper_id = paper.attrib["id"]
                title = "".join(paper.find("title").itertext())
                uid = f"{collection_id}-{volume_id}.{paper_id}"
                authors = [
                    " ".join(author.itertext()) for author in paper.findall("author")
                ]
                authors = "|".join(authors)

                if paper.find("abstract") is not None:
                    abstract = "".join(paper.find("abstract").itertext())
                else:
                    abstract = ""

                link = f"https://www.aclweb.org/anthology/{uid}"

                track = mapping[venue]
                kind = None

                if track.startswith("W"):
                    kind = "workshop"
                elif track == "main":
                    kind = "long"
                else:
                    kind = "findings"

                assert kind

                paper = Paper(
                    uid=uid,
                    ws_id=ws_id,
                    title=title,
                    authors=authors,
                    abstract=abstract,
                    track=track,
                    kind=kind,
                    link=link,
                )

                papers.append(paper)

    return papers


def read_wmt_bib() -> List[Paper]:
    result = []
    with open("downloads/2020.wmt-1.0.bib") as f:
        bib = database.parse_file(f)

        for i, entry in enumerate(bib.entries.values()):
            if entry.type == "book":
                continue

            title = LatexNodes2Text().latex_to_text(entry.fields["title"])
            url = entry.fields["url"]
            abstract = LatexNodes2Text().latex_to_text(entry.fields["abstract"])
            author = "|".join(
                [
                    " ".join(reversed(str(e).split(", ")))
                    for e in entry.persons["author"]
                ]
            )

            uid = url.replace("https://www.aclweb.org/anthology/", "")
            url = "https://www.statmt.org/wmt20/pdf/" + uid + ".pdf"

            paper = Paper(
                uid=f"WS-2.{uid}",
                ws_id="WS-2",
                title=title,
                authors=author,
                abstract=abstract,
                track="WS-2",
                kind="workshop",
                link=url,
            )
            result.append(paper)
    return result


def is_not_paper(row) -> bool:
    uid = row["Unique ID"].lower().strip()
    title = row["Title"].lower().strip()

    return (
        ("invited" in uid)
        or ("challenge" in uid)
        or ("invited" in title)
        or ("keynote" in title)
        or ("keynote" in uid)
        or (row["Unique ID"] == "Shared task")
        or (title == "tba" and "paper" not in uid)
    )


def add_invited_talks(slideslive: pd.DataFrame):
    talks_per_workshop = defaultdict(list)
    fixed_workshop_titles_df = pd.read_excel(PATH_WORKSHOP_TALKS)

    id_to_fixed_talk = {
        f"{row['Unique ID']}.{row['Organizer track name']}".strip(): (
            row["Title"],
            row["Speakers"],
        )
        for _, row in fixed_workshop_titles_df.iterrows()
    }

    for _, row in slideslive.iterrows():
        if not is_not_paper(row):
            continue

        uid = f"{row['Unique ID']}.{row['Organizer track name']}".strip()

        if uid in id_to_fixed_talk:
            title, speakers = id_to_fixed_talk[uid]
        else:
            title = row["Title"].strip()
            speakers = row["Speakers"].strip()

        presentation_id = row["SlidesLive link"].replace("https://slideslive.com/", "")

        if presentation_id == "38939447":
            print("removing", presentation_id)
            continue

        talks_per_workshop[row["Organizer track name"].strip()].append(
            {"title": title, "speakers": speakers, "presentation_id": presentation_id}
        )

    return talks_per_workshop


def get_zooms() -> Dict[str, List[str]]:
    df = pd.read_excel(PATH_ZOOM_ACCOUNTS_WITH_PASSWORDS, sheet_name="Workshops")

    mapping = {
        "5th Conference on Machine Translation (WMT20)": "Fifth Conference on Machine Translation (WMT20)",
        "Search-Oriented Conversational AI (SCAI)": "Search-Oriented Conversational AI (SCAI) 2",
        "4th Workshop on Online Abuse and Harms (WOAH) a.k.a. ALW": "The Fourth Workshop on Online Abuse and Harms (WOAH) a.k.a. ALW",
        "Evaluation and Comparison of NLP Systems (Eval4NLP)": "Evaluation and Comparison of NLP Systems",
        "2nd Workshop for NLP Open Source Software (NLP-OSS)": "Second Workshop for NLP Open Source Software (NLP-OSS)",
    }

    zooms = defaultdict(list)
    for _, row in df.iterrows():
        uid = row["WS Name"]
        uid = mapping.get(uid, uid)
        zooms[uid].append(row["Personal Meeting LINK"])

        for i in range(row["# of accounts"] - 1):
            zooms[uid].append(row[f"Personal Meeting LINK.{i+1}"])

    return zooms


if __name__ == "__main__":
    # download_slideslive()
    download_workshops()
    # download_zooms()

    # load_csv()
    data = build_workshops_basics()
    slideslive = load_slideslive()
    generate_workshop_papers(slideslive)
    talks = add_invited_talks(slideslive)

    fix_talks = slideslive[[is_not_paper(r) for _, r in slideslive.iterrows()]]
    fix_talks.to_csv(
        "yamls/fix_talks.csv",
        index=False,
        columns=["Organizer track name", "Unique ID", "Title", "Speakers"],
    )

    for ws in data:
        uid = ws["UID"]
        ws["prerecorded_talks"] = talks[uid]

    yaml.scalarstring.walk_tree(data)

    with open(PATH_YAMLS / "workshops.yml", "w") as f:
        yaml.dump(data, f, Dumper=ruamel.yaml.RoundTripDumper)
