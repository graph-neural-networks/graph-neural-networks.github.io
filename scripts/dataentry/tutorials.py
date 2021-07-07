import csv
import random
from collections import defaultdict
from dataclasses import dataclass
from typing import List, Dict

import pytz
import ruamel
from openpyxl import load_workbook
from ruamel import yaml

from datetime import datetime

import pandas as pd

from scripts.dataentry.paths import *


@dataclass
class Session:
    name: str
    start: datetime
    end: datetime
    host: str
    zoom_link: str = None


@dataclass
class Tutorial:
    uid: str
    sessions: List[Session]
    abstract: str
    website: str
    material: str
    slides: str


def load_schedule() -> Dict[str, Tutorial]:
    wb = load_workbook(PATH_TUTORIALS_SCHEDULE)
    zooms = load_zooms()

    data = {}
    for ws in wb.worksheets[1:]:

        tutorial_id = ws["B2"].value or ""
        abstract = ws["B3"].value or ""
        website = ws["B4"].value or ""
        material = ws["B5"].value or ""
        slides = ws["B6"].value or ""

        tutorial_id = tutorial_id.strip()
        abstract = abstract.strip()
        website = website.strip()
        material = material.strip()
        slides = slides.strip()

        ws.delete_rows(1, 9)
        ws.delete_cols(7, 100)
        df = pd.DataFrame(
            ws.values,
            columns=[
                "Session Name",
                "Day",
                "Start Time",
                "End Time",
                "Time Zone",
                "Host",
            ],
        )
        df.dropna(subset=["Session Name"], inplace=True)

        sessions = []
        for idx, row in df.iterrows():
            name = row["Session Name"].strip()
            host = row["Host"] or None

            day = row["Day"]
            start_time = row["Start Time"]

            if isinstance(start_time, datetime):
                start_time = start_time.time()

            end_time = row["End Time"]
            assert row["Time Zone"] == "UTC-0", "Was" + str(row["Time Zone"])

            tz = pytz.utc

            start = datetime.combine(day.date(), start_time)
            start = tz.localize(start)

            end = datetime.combine(day.date(), end_time)
            end = tz.localize(end)

            session = Session(name, start, end, host, zooms[tutorial_id])
            sessions.append(session)

        tutorial = Tutorial(
            tutorial_id,
            sessions,
            abstract=abstract,
            website=website,
            material=material,
            slides=slides,
        )
        assert tutorial_id not in data
        data[tutorial_id] = tutorial

    return data


def load_slideslive() -> pd.DataFrame:
    return pd.read_csv(PATH_SLIDESLIVE_OTHER)


def load_zooms() -> Dict[str, str]:
    df = pd.read_excel(PATH_ZOOM_ACCOUNTS_WITH_PASSWORDS, sheet_name="Tutorials")
    result = {}

    for _, row in df.iterrows():
        number = row["UID"].split(".")[-1]
        uid = "T" + number
        result[uid] = row["Personal Meeting LINK"]

    return result


def generate_yaml():
    tutorials = load_schedule()
    tutorials_overview = {
        row["tutorialNumber"]: row
        for _, row in pd.read_csv(PATH_TUTORIALS_OVERVIEW).iterrows()
    }
    slideslive = load_slideslive()
    id_to_recording = {}
    for _, row in slideslive.iterrows():
        if row["Organizer track name"] == "Demo":
            title = row["Title"]
            uid, title = title.split(":", 1)
            uid = uid.strip()
            id_to_recording[uid] = row["SlidesLive link"].replace(
                "https://slideslive.com/", ""
            )

    data = []

    info = """
    This tutorial has a prerecorded talk on this page (see below) that you can watch anytime during the conference. It
    also has two live sessions that will be conducted on Zoom and will be livestreamed on this page. Additionally, it has
    a chat window that you can use to have discussions with the tutorial teachers and other attendees anytime during the
    conference.""".strip()

    info_live = """
    This tutorial has slides that you see can anytime (It does not have any
    prerecorded talk). It will be conducted entirely live on Zoom and will be
    livestreamed on this page. It has a chat window that you can use to have
    discussions with the tutorial teachers and other attendees anytime during the
    conference.""".strip()

    channels = {
        "T1": "tutorial-1-interpreting-predictions",
        "T2": "tutorial-2-fact-checking",
        "T3": "tutorial-3-high-performance",
        "T4": "tutorial-4-machine-reasoning",
        "T5": "tutorial-5-spatial-language",
        "T6": "tutorial-6-simultaneous-translation",
        "T7": "tutorial-7-world-of-generation",
    }

    for tutorial in tutorials.values():
        entry = {
            "UID": tutorial.uid,
            "title": tutorials_overview[tutorial.uid]["title"],
            "organizers": tutorials_overview[tutorial.uid]["authors"],
            "abstract": tutorial.abstract,
            "info": info if tutorial.uid != "T1" else info_live,
            "rocketchat_channel": channels[tutorial.uid],
            "sessions": [],
        }

        if tutorial.uid in id_to_recording:
            entry["prerecorded"] = id_to_recording[tutorial.uid]

        if tutorial.website:
            entry["website"] = tutorial.website

        if tutorial.material:
            entry["material"] = tutorial.material

        if tutorial.slides:
            entry["slides"] = tutorial.slides

        for session in tutorial.sessions:
            e = {
                "name": session.name,
                "start_time": session.start,
                "end_time": session.end,
                "zoom_link": session.zoom_link,
            }
            if session.host:
                e["hosts"] = session.host

            entry["sessions"].append(e)

        data.append(entry)
    yaml.scalarstring.walk_tree(data)

    with open(PATH_YAMLS / "tutorials.yml", "w") as f:
        yaml.dump(data, f, Dumper=ruamel.yaml.RoundTripDumper)


if __name__ == "__main__":
    download_zooms()
    download_slideslive()
    download_tutorials()
    generate_yaml()
