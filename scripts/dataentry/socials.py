import csv
import random
from collections import defaultdict
from dataclasses import dataclass
from typing import List, Dict

import pytz
import ruamel
from openpyxl import load_workbook
from ruamel import yaml

from datetime import datetime

import pandas as pd

# https://docs.google.com/spreadsheets/d/1IDk3K1JD1hvH_hvyMy6TeRuE2F6DQDfpgwNpTIP9KgI/edit#gid=1276180010
from scripts.dataentry.paths import *


def generate_socials():
    wb = load_workbook(PATH_SOCIALS)

    ws = wb.worksheets[0]
    ws.delete_cols(10, 100)
    df = pd.DataFrame(
        ws.values,
        columns=[
            "ID",
            "Event Type",
            "Event",
            "Pre-recorded/live",
            "Platform",
            "Organizers",
            "Contact person",
            "Email address",
            "Channel Name",
        ],
    )
    df = df.dropna(subset=["ID"])
    df = df.drop([df.index[-2]])

    zoom_df = pd.read_excel(
        PATH_ZOOM_ACCOUNTS_WITH_PASSWORDS, sheet_name="Affinity"
    ).fillna("")

    zooms = {}
    for _, row in zoom_df.iterrows():
        number = row["UID"].split(".")[-1]
        uid = row["UID"][0].upper() + number

        assert uid not in zooms

        link = row["Personal Meeting LINK"]
        if link:
            zooms[uid] = link

    id_to_organizers = {
        row["ID"]: [e.strip() for e in row["Organizers"].split(",")]
        for _, row in df.iterrows()
        if row["Organizers"]
    }
    id_to_name = {row["ID"]: row["Event"] for _, row in df.iterrows()}
    id_to_channel = {row["ID"]: row["Channel Name"] for _, row in df.iterrows()}
    id_to_location = {row["ID"]: row["Platform"] for _, row in df.iterrows()}

    result = []
    for ws in wb.worksheets[2:]:
        data = {}

        uid = ws["B2"].value
        description = ws["B3"].value
        website = ws["B4"].value
        name = id_to_name[uid]

        # print(uid)

        images = {
            "A1": "static/images/socials/queer_in_ai.png",
            "A2": "static/images/socials/VegNLP-logo.png",
            "A3": "static/images/socials/LXAI-navlogo.png",
            "A4": "static/images/socials/NorthAfricansInNLP.png",
        }

        if uid != "M2":
            ws.delete_rows(1, 9)
            ws.delete_cols(7, 100)
            df = pd.DataFrame(
                ws.values,
                columns=[
                    "Session Name",
                    "Day",
                    "Start Time",
                    "End Time",
                    "Time Zone",
                    "Host",
                ],
            )
        else:
            ws.delete_rows(1, 9)
            ws.delete_cols(8, 100)
            df = pd.DataFrame(
                ws.values,
                columns=[
                    "Session Name",
                    "Day",
                    "Start Time",
                    "End Time",
                    "Time Zone",
                    "Host",
                    "Zoom Link",
                ],
            )
        df.dropna(subset=["Session Name"], inplace=True)

        data["name"] = name
        data["description"] = description
        data["UID"] = uid
        data["organizers"] = {"members": id_to_organizers[uid]}

        if uid in images:
            data["image"] = images[uid]

        if uid.startswith("B"):
            data["image"] = "static/images/emnlp2020/acl-logo.png"

        if website:
            data["website"] = website

        data["rocketchat_channel"] = id_to_channel[uid]
        data["location"] = id_to_location[uid]

        if uid in zooms:
            assert "Zoom" in data["location"], data["location"]
            data["zoom_link"] = zooms[uid]

        result.append(data)

        sessions = []
        for idx, row in df.iterrows():
            name = "S-" + row["Session Name"].strip()

            if (uid.startswith("B") or uid.startswith("M")) and row["Host"]:
                name = name + " with " + row["Host"]

            day = row["Day"]
            start_time = row["Start Time"]

            if isinstance(start_time, datetime):
                start_time = start_time.time()

            end_time = row["End Time"]
            # assert row["Time Zone"] == "UTC-0", "Was" + str(row["Time Zone"] )

            tz = pytz.utc

            start = datetime.combine(day.date(), start_time)
            start = tz.localize(start)

            end = datetime.combine(day.date(), end_time)
            end = tz.localize(end)

            e = {
                "name": name,
                "start_time": start,
                "end_time": end,
            }
            # print(df.columns)
            if "Zoom Link" in df.columns.values:
                e["link"] = row["Zoom Link"]

            sessions.append(e)

        data["sessions"] = sessions

    result.sort(key=lambda x: x["UID"])
    yaml.scalarstring.walk_tree(result)

    with open("yamls/socials.yml", "w") as f:
        yaml.dump(result, f, Dumper=ruamel.yaml.RoundTripDumper)


if __name__ == "__main__":
    download_zooms()
    download_socials()
    generate_socials()
